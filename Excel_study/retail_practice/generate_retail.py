#!/usr/bin/env python3
"""小売データ分析演習 Excel 生成スクリプト"""

import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

random.seed(42)

# ──────────────────────────────
# 1. マスタ定義
# ──────────────────────────────
CATEGORIES_PRODUCTS = {
    "家電":   {"スマートフォン":89800,"タブレット":65000,"ノートPC":128000,"イヤホン":15800,"充電器":2980},
    "衣料品": {"Tシャツ":3500,"ジャケット":18000,"パンツ":7800,"スニーカー":12000,"バッグ":25000},
    "食品":   {"コーヒー":1200,"紅茶":800,"チョコレート":500,"クッキー":450,"ナッツ":980},
    "日用品": {"シャンプー":1500,"洗剤":800,"タオル":1200,"歯ブラシ":350,"ティッシュ":600},
    "スポーツ":{"ヨガマット":4500,"ダンベル":8000,"プロテイン":5800,"ランニングシューズ":14000,"スポーツウェア":6500},
    "書籍":   {"ビジネス書":1800,"小説":1500,"参考書":2500,"雑誌":780,"漫画":500},
}
REGIONS       = ["東京","大阪","名古屋","福岡","札幌"]
PAYMENTS      = ["クレジットカード","現金","電子マネー","QRコード決済"]
RANKS         = ["ゴールド","シルバー","ブロンズ","一般"]
CAT_W         = [0.25, 0.20, 0.20, 0.15, 0.10, 0.10]
REG_W         = [0.35, 0.25, 0.15, 0.15, 0.10]
RANK_W        = [0.10, 0.20, 0.30, 0.40]

# ──────────────────────────────
# 2. データ生成（1,000件）
# ──────────────────────────────
N = 1000
rows = []
start = date(2025, 1, 1)

for i in range(N):
    d   = start + timedelta(days=random.randint(0, 364))
    cat = random.choices(list(CATEGORIES_PRODUCTS), weights=CAT_W)[0]
    prd = random.choice(list(CATEGORIES_PRODUCTS[cat]))
    up  = CATEGORIES_PRODUCTS[cat][prd]
    qty = random.randint(1, 5)
    rows.append([
        f"TRX-{i+1:05d}",
        d,
        f"CUS-{random.randint(1,300):03d}",
        random.choices(RANKS,  weights=RANK_W)[0],
        random.choices(REGIONS, weights=REG_W)[0],
        cat, prd, up, qty, up * qty,
        random.choice(PAYMENTS),
    ])

rows.sort(key=lambda x: x[1])   # 日付順

DATA_HEADERS = ["取引ID","取引日","顧客ID","会員ランク","地域",
                "カテゴリ","商品名","単価","数量","売上金額","支払方法"]

# ──────────────────────────────
# 3. スタイル定義
# ──────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
SUB_FONT  = Font(bold=True, color="FFFFFF", size=10)
Q_FILL    = PatternFill("solid", fgColor="FFF2CC")
Q_FONT    = Font(bold=True, color="7F6000", size=10)
ANS_FILL  = PatternFill("solid", fgColor="E2EFDA")
ANS_FONT  = Font(bold=True, color="375623", size=10)
GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_AL   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT_AL  = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_border():
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_hdr(cell, text, fill=None, font=None, align=None):
    cell.value     = text
    cell.fill      = fill  or HDR_FILL
    cell.font      = font  or HDR_FONT
    cell.alignment = align or CENTER
    cell.border    = hdr_border()

def apply_body(cell, value=None, fmt=None, align=None):
    if value is not None:
        cell.value = value
    if fmt:
        cell.number_format = fmt
    cell.alignment = align or CENTER
    cell.border    = thin_border()

# ──────────────────────────────
# 4. ワークブック作成
# ──────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# =============================
# Sheet 0: README
# =============================
ws0 = wb.create_sheet("0_README")
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 60
ws0.column_dimensions["B"].width = 40

readme = [
    ("小売データ分析演習ブック", ""),
    ("", ""),
    ("■ 構成", ""),
    ("シート名","内容"),
    ("1_売上データ", "1,000件の取引データ（2025年通年）"),
    ("2_商品マスタ", "商品コード・名称・カテゴリ・標準単価"),
    ("3_問題", "データ分析の演習問題（解答セル空白）"),
    ("4_解答", "数式つき解答（カンニング用）"),
    ("", ""),
    ("■ データ仕様", ""),
    ("顧客ID", "CUS-001 〜 CUS-300（架空）"),
    ("期間","2025-01-01 〜 2025-12-31"),
    ("カテゴリ","家電 / 衣料品 / 食品 / 日用品 / スポーツ / 書籍"),
    ("地域","東京 / 大阪 / 名古屋 / 福岡 / 札幌"),
    ("会員ランク","ゴールド / シルバー / ブロンズ / 一般"),
    ("支払方法","クレジットカード / 現金 / 電子マネー / QRコード決済"),
]

for r, (a, b) in enumerate(readme, 1):
    ca, cb = ws0.cell(r,1), ws0.cell(r,2)
    if r == 1:
        ca.value = a; ca.font = Font(bold=True,size=14,color="1F4E79"); ca.alignment = LEFT_AL
    elif a == "シート名":
        apply_hdr(ca,a); apply_hdr(cb,b)
    elif a.startswith("■"):
        ca.value = a; ca.font = Font(bold=True,color="2E75B6",size=11); ca.alignment = LEFT_AL
    else:
        ca.value = a; ca.alignment = LEFT_AL; ca.border = thin_border()
        cb.value = b; cb.alignment = LEFT_AL; cb.border = thin_border()

# =============================
# Sheet 1: 売上データ
# =============================
ws1 = wb.create_sheet("1_売上データ")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A2"

col_widths = [14,13,11,11,9,9,20,10,7,13,16]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

for ci, h in enumerate(DATA_HEADERS, 1):
    apply_hdr(ws1.cell(1, ci), h)

DATE_FMT   = "yyyy/mm/dd"
NUM_FMT    = "#,##0"
for ri, row in enumerate(rows, 2):
    alt_fill = PatternFill("solid", fgColor="EBF3FB") if ri % 2 == 0 else None
    for ci, val in enumerate(row, 1):
        cell = ws1.cell(ri, ci, val)
        cell.border    = thin_border()
        cell.alignment = CENTER
        if alt_fill:
            cell.fill = alt_fill
        if ci == 2:
            cell.number_format = DATE_FMT
        if ci in (8, 10):
            cell.number_format = NUM_FMT

ws1.auto_filter.ref = f"A1:{get_column_letter(len(DATA_HEADERS))}1"

# =============================
# Sheet 2: 商品マスタ
# =============================
ws2 = wb.create_sheet("2_商品マスタ")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 12

for ci, h in enumerate(["商品コード","商品名","カテゴリ","標準単価"], 1):
    apply_hdr(ws2.cell(1, ci), h)

prd_code = 1
master_rows = []
for cat, prods in CATEGORIES_PRODUCTS.items():
    for pname, price in prods.items():
        master_rows.append([f"P-{prd_code:03d}", pname, cat, price])
        prd_code += 1

for ri, (code, pname, cat, price) in enumerate(master_rows, 2):
    alt = PatternFill("solid", fgColor="EBF3FB") if ri % 2 == 0 else None
    for ci, v in enumerate([code, pname, cat, price], 1):
        cell = ws2.cell(ri, ci, v)
        cell.border    = thin_border()
        cell.alignment = CENTER
        if alt: cell.fill = alt
        if ci == 4: cell.number_format = NUM_FMT

# =============================
# Sheet 3: 問題
# =============================
ws3 = wb.create_sheet("3_問題")
ws3.sheet_view.showGridLines = False

def q_title(ws, row, text):
    cell = ws.cell(row, 1, text)
    cell.fill      = Q_FILL
    cell.font      = Q_FONT
    cell.alignment = LEFT_AL
    cell.border    = hdr_border()
    ws.merge_cells(f"A{row}:J{row}")

def q_label(ws, row, col, text):
    cell = ws.cell(row, col, text)
    cell.fill = GRAY_FILL
    cell.font = Font(bold=True, size=10)
    cell.alignment = CENTER
    cell.border = thin_border()

def blank_cell(ws, row, col, w=12):
    cell = ws.cell(row, col)
    cell.fill = PatternFill("solid", fgColor="FFFDE7")
    cell.border = thin_border()
    cell.alignment = CENTER

# 列幅設定
for ci, w in enumerate([18,12,12,12,12,12,12,12,12,12], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

cur_row = 1

# Q1 月別売上合計
q_title(ws3, cur_row, "Q1｜月別売上合計（SUMPRODUCT + MONTH関数）")
cur_row += 1
for ci, h in enumerate(["月","売上合計","前月比(%)"], 1):
    q_label(ws3, cur_row, ci, h)
cur_row += 1
for m in range(1, 13):
    ws3.cell(cur_row, 1, m).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    blank_cell(ws3, cur_row, 2)
    blank_cell(ws3, cur_row, 3)
    cur_row += 1

cur_row += 1

# Q2 カテゴリ別売上
q_title(ws3, cur_row, "Q2｜カテゴリ別売上合計・構成比（SUMIF / SUMPRODUCT）")
cur_row += 1
for ci, h in enumerate(["カテゴリ","売上合計","構成比(%)"], 1):
    q_label(ws3, cur_row, ci, h)
cur_row += 1
for cat in CATEGORIES_PRODUCTS:
    ws3.cell(cur_row, 1, cat).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    blank_cell(ws3, cur_row, 2)
    blank_cell(ws3, cur_row, 3)
    cur_row += 1

cur_row += 1

# Q3 地域×カテゴリ クロス集計
q_title(ws3, cur_row, "Q3｜地域×カテゴリ クロス集計（SUMIFS）")
cur_row += 1
cats = list(CATEGORIES_PRODUCTS.keys())
q_label(ws3, cur_row, 1, "地域＼カテゴリ")
for ci, cat in enumerate(cats, 2):
    q_label(ws3, cur_row, ci, cat)
cur_row += 1
for reg in REGIONS:
    ws3.cell(cur_row, 1, reg).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    ws3.cell(cur_row, 1).font = Font(bold=True, size=10)
    for ci in range(2, len(cats)+2):
        blank_cell(ws3, cur_row, ci)
    cur_row += 1

cur_row += 1

# Q4 会員ランク別平均単価
q_title(ws3, cur_row, "Q4｜会員ランク別 平均購入金額・件数（AVERAGEIF / COUNTIF）")
cur_row += 1
for ci, h in enumerate(["会員ランク","平均購入金額","取引件数","合計金額"], 1):
    q_label(ws3, cur_row, ci, h)
cur_row += 1
for rank in RANKS:
    ws3.cell(cur_row, 1, rank).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    for ci in range(2, 5):
        blank_cell(ws3, cur_row, ci)
    cur_row += 1

cur_row += 1

# Q5 支払方法別件数・売上
q_title(ws3, cur_row, "Q5｜支払方法別 件数・売上合計（COUNTIF / SUMIF）")
cur_row += 1
for ci, h in enumerate(["支払方法","件数","売上合計","平均単価"], 1):
    q_label(ws3, cur_row, ci, h)
cur_row += 1
for pay in PAYMENTS:
    ws3.cell(cur_row, 1, pay).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    for ci in range(2, 5):
        blank_cell(ws3, cur_row, ci)
    cur_row += 1

cur_row += 1

# Q6 複合条件 SUMIFS
q_title(ws3, cur_row, "Q6｜複合条件集計（SUMIFS / COUNTIFS）")
cur_row += 1
questions6 = [
    "① ゴールド会員 × 東京 の売上合計",
    "② 家電 × クレジットカード の件数",
    "③ 2025年Q2（4〜6月）の売上合計",
    "④ 売上金額 10,000円超 の取引件数",
    "⑤ 大阪 × ブロンズ会員 × 食品 の売上合計",
]
for q_text in questions6:
    ws3.cell(cur_row, 1, q_text).alignment = LEFT_AL
    ws3.cell(cur_row, 1).border = thin_border()
    ws3.merge_cells(f"A{cur_row}:D{cur_row}")
    blank_cell(ws3, cur_row, 5)
    ws3.merge_cells(f"E{cur_row}:H{cur_row}")
    cur_row += 1

cur_row += 1

# Q7 VLOOKUP
q_title(ws3, cur_row, "Q7｜商品マスタ参照（VLOOKUP / XLOOKUP / IFERROR）")
cur_row += 1
for ci, h in enumerate(["商品コード","商品名（VLOOKUP）","カテゴリ（XLOOKUP）","標準単価（IFERROR）"], 1):
    q_label(ws3, cur_row, ci, h)
cur_row += 1
sample_codes = ["P-001","P-007","P-015","P-025","P-999"]  # P-999 は存在しない
for code in sample_codes:
    ws3.cell(cur_row, 1, code).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    for ci in range(2, 5):
        blank_cell(ws3, cur_row, ci)
    cur_row += 1

cur_row += 1

# Q8 ランキング
q_title(ws3, cur_row, "Q8｜売上ランキング 上位10（LARGE / RANK / INDEX+MATCH）")
cur_row += 1
for ci, h in enumerate(["順位","最大売上金額（LARGE）","その取引ID（INDEX+MATCH）"], 1):
    q_label(ws3, cur_row, ci, h)
cur_row += 1
for rank_n in range(1, 11):
    ws3.cell(cur_row, 1, rank_n).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    blank_cell(ws3, cur_row, 2)
    blank_cell(ws3, cur_row, 3)
    cur_row += 1

cur_row += 1

# Q9 顧客別集計
q_title(ws3, cur_row, "Q9｜顧客別 購入回数・総購入額（COUNTIF / SUMIF）")
cur_row += 1
for ci, h in enumerate(["顧客ID","購入回数","総購入額","平均購入単価"], 1):
    q_label(ws3, cur_row, ci, h)
cur_row += 1
sample_customers = ["CUS-001","CUS-010","CUS-050","CUS-100","CUS-200"]
for cid in sample_customers:
    ws3.cell(cur_row, 1, cid).border = thin_border()
    ws3.cell(cur_row, 1).alignment = CENTER
    for ci in range(2, 5):
        blank_cell(ws3, cur_row, ci)
    cur_row += 1

cur_row += 1

# Q10 統計分析
q_title(ws3, cur_row, "Q10｜統計分析（MEDIAN / STDEV / PERCENTILE / MAX / MIN）")
cur_row += 1
stats_items = [
    "売上金額 の中央値（MEDIAN）",
    "売上金額 の標準偏差（STDEV）",
    "売上金額 の90パーセンタイル（PERCENTILE）",
    "1取引あたり最大売上金額（MAX）",
    "1取引あたり最小売上金額（MIN）",
    "売上金額 の四分位範囲（Q3-Q1）",
]
for item in stats_items:
    ws3.cell(cur_row, 1, item).alignment = LEFT_AL
    ws3.cell(cur_row, 1).border = thin_border()
    ws3.merge_cells(f"A{cur_row}:D{cur_row}")
    blank_cell(ws3, cur_row, 5)
    ws3.merge_cells(f"E{cur_row}:H{cur_row}")
    cur_row += 1

# =============================
# Sheet 4: 解答
# =============================
ws4 = wb.create_sheet("4_解答")
ws4.sheet_view.showGridLines = False

for ci, w in enumerate([28,30,20,20,20,20,20,20,20,20], 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w

DATA = "'1_売上データ'"
MSTR = "'2_商品マスタ'"

def ans_title(ws, row, text):
    cell = ws.cell(row, 1, text)
    cell.fill      = ANS_FILL
    cell.font      = ANS_FONT
    cell.alignment = LEFT_AL
    cell.border    = hdr_border()
    ws.merge_cells(f"A{row}:J{row}")

def ans_row(ws, row, label, formula, note=""):
    lc = ws.cell(row, 1, label);  lc.border=thin_border(); lc.alignment=LEFT_AL; lc.fill=GRAY_FILL
    fc = ws.cell(row, 2, formula); fc.border=thin_border(); fc.alignment=LEFT_AL
    fc.font = Font(name="Courier New", size=9, color="1F4E79")
    if note:
        nc = ws.cell(row, 3, note); nc.border=thin_border(); nc.alignment=LEFT_AL
        nc.font = Font(size=9, color="595959", italic=True)

ar = 1

# A1 月別売上
ans_title(ws4, ar, "Q1｜月別売上合計"); ar+=1
ans_row(ws4, ar, "B2（1月売上）",
    f"=SUMPRODUCT((MONTH({DATA}!$B$2:$B$1001)=A2)*{DATA}!$J$2:$J$1001)",
    "A列に月番号(1〜12)が入っている前提。MONTH()で月を抽出し合計")
ar+=1
ans_row(ws4, ar, "B3以降",
    "→ B2をB13までコピー（A列の月番号が自動的に変わる）", ""); ar+=1
ans_row(ws4, ar, "C3（前月比）",
    "=IFERROR(B3/B2-1,\"\")",
    "書式をパーセント表示に。1月(C2)はIFERRORで空白"); ar+=2

# A2 カテゴリ別
ans_title(ws4, ar, "Q2｜カテゴリ別売上合計・構成比"); ar+=1
ans_row(ws4, ar, "B2（売上合計）",
    f"=SUMIF({DATA}!$F:$F,A2,{DATA}!$J:$J)",
    "A列にカテゴリ名が入っている前提"); ar+=1
ans_row(ws4, ar, "C2（構成比）",
    "=B2/SUM($B$2:$B$7)",
    "書式をパーセント(小数2桁)に。$B$2:$B$7は全カテゴリ合計セル範囲"); ar+=2

# A3 クロス集計
ans_title(ws4, ar, "Q3｜地域×カテゴリ クロス集計"); ar+=1
ans_row(ws4, ar, "B3（東京×家電）",
    f"=SUMIFS({DATA}!$J:$J,{DATA}!$E:$E,$A3,{DATA}!$F:$F,B$2)",
    "行ヘッダ($A3)と列ヘッダ(B$2)を混合参照にして全セルにコピー可能"); ar+=2

# A4 会員ランク別
ans_title(ws4, ar, "Q4｜会員ランク別 平均・件数・合計"); ar+=1
ans_row(ws4, ar, "B2（平均購入金額）",
    f"=AVERAGEIF({DATA}!$D:$D,A2,{DATA}!$J:$J)",
    "書式を #,##0 に"); ar+=1
ans_row(ws4, ar, "C2（取引件数）",
    f"=COUNTIF({DATA}!$D:$D,A2)", ""); ar+=1
ans_row(ws4, ar, "D2（合計金額）",
    f"=SUMIF({DATA}!$D:$D,A2,{DATA}!$J:$J)", ""); ar+=2

# A5 支払方法別
ans_title(ws4, ar, "Q5｜支払方法別 件数・売上合計・平均"); ar+=1
ans_row(ws4, ar, "B2（件数）",
    f"=COUNTIF({DATA}!$K:$K,A2)", ""); ar+=1
ans_row(ws4, ar, "C2（売上合計）",
    f"=SUMIF({DATA}!$K:$K,A2,{DATA}!$J:$J)", ""); ar+=1
ans_row(ws4, ar, "D2（平均単価）",
    "=IFERROR(C2/B2,\"\")", "件数0の場合エラー回避"); ar+=2

# A6 複合条件
ans_title(ws4, ar, "Q6｜複合条件集計（SUMIFS / COUNTIFS）"); ar+=1
ans_row(ws4, ar, "① ゴールド×東京 売上",
    f'=SUMIFS({DATA}!$J:$J,{DATA}!$D:$D,"ゴールド",{DATA}!$E:$E,"東京")', ""); ar+=1
ans_row(ws4, ar, "② 家電×クレカ 件数",
    f'=COUNTIFS({DATA}!$F:$F,"家電",{DATA}!$K:$K,"クレジットカード")', ""); ar+=1
ans_row(ws4, ar, "③ Q2（4〜6月）売上",
    f'=SUMPRODUCT((MONTH({DATA}!$B$2:$B$1001)>=4)*(MONTH({DATA}!$B$2:$B$1001)<=6)*{DATA}!$J$2:$J$1001)',
    "SUMIFS+DATEで代替可: =SUMIFS(J:J,B:B,\">=\"&DATE(2025,4,1),B:B,\"<\"&DATE(2025,7,1))"); ar+=1
ans_row(ws4, ar, "④ 10,000円超 件数",
    f'=COUNTIF({DATA}!$J:$J,">10000")', "条件に演算子を含む場合は文字列で"); ar+=1
ans_row(ws4, ar, "⑤ 大阪×ブロンズ×食品 売上",
    f'=SUMIFS({DATA}!$J:$J,{DATA}!$E:$E,"大阪",{DATA}!$D:$D,"ブロンズ",{DATA}!$F:$F,"食品")', ""); ar+=2

# A7 VLOOKUP
ans_title(ws4, ar, "Q7｜商品マスタ参照"); ar+=1
ans_row(ws4, ar, "B2（VLOOKUP 商品名）",
    f"=VLOOKUP(A2,{MSTR}!$A:$D,2,FALSE)",
    "第4引数FALSE=完全一致。列番号2=商品名"); ar+=1
ans_row(ws4, ar, "C2（XLOOKUP カテゴリ）",
    f'=XLOOKUP(A2,{MSTR}!$A:$A,{MSTR}!$C:$C,"該当なし")',
    "Excel365/2021以降。見つからない場合は第4引数で表示"); ar+=1
ans_row(ws4, ar, "D2（IFERROR 標準単価）",
    f'=IFERROR(VLOOKUP(A2,{MSTR}!$A:$D,4,FALSE),"未登録")',
    "存在しないコードのとき #N/A→「未登録」に置換"); ar+=2

# A8 ランキング
ans_title(ws4, ar, "Q8｜売上ランキング 上位10"); ar+=1
ans_row(ws4, ar, "B2（LARGE）",
    f"=LARGE({DATA}!$J:$J,A2)",
    "A列に順位(1〜10)が入っている前提"); ar+=1
ans_row(ws4, ar, "C2（INDEX+MATCH）",
    f"=INDEX({DATA}!$A:$A,MATCH(B2,{DATA}!$J:$J,0))",
    "同額が複数ある場合は最初にヒットした取引IDを返す"); ar+=2

# A9 顧客別
ans_title(ws4, ar, "Q9｜顧客別 購入回数・総購入額"); ar+=1
ans_row(ws4, ar, "B2（購入回数）",
    f"=COUNTIF({DATA}!$C:$C,A2)", ""); ar+=1
ans_row(ws4, ar, "C2（総購入額）",
    f"=SUMIF({DATA}!$C:$C,A2,{DATA}!$J:$J)", ""); ar+=1
ans_row(ws4, ar, "D2（平均購入単価）",
    "=IFERROR(C2/B2,\"\")", ""); ar+=2

# A10 統計
ans_title(ws4, ar, "Q10｜統計分析"); ar+=1
ans_row(ws4, ar, "売上金額 中央値",
    f"=MEDIAN({DATA}!$J:$J)", ""); ar+=1
ans_row(ws4, ar, "売上金額 標準偏差",
    f"=STDEV({DATA}!$J$2:$J$1001)",
    "STDEVは標本標準偏差（n-1）。母標準偏差はSTDEVP"); ar+=1
ans_row(ws4, ar, "90パーセンタイル",
    f"=PERCENTILE({DATA}!$J$2:$J$1001,0.9)",
    "0〜1で指定。0.9=上位10%の閾値"); ar+=1
ans_row(ws4, ar, "MAX",
    f"=MAX({DATA}!$J:$J)", ""); ar+=1
ans_row(ws4, ar, "MIN",
    f"=MIN({DATA}!$J$2:$J$1001)", "0を除きたい場合=MINIFS(J:J,J:J,\">0\")"); ar+=1
ans_row(ws4, ar, "四分位範囲（IQR）",
    f"=PERCENTILE({DATA}!$J$2:$J$1001,0.75)-PERCENTILE({DATA}!$J$2:$J$1001,0.25)",
    "Q3(75%ile) - Q1(25%ile)"); ar+=2

# =============================
# Save
# =============================
OUT = "/Users/kengo.nomura/Downloads/excel/retail_practice/小売データ分析演習.xlsx"
wb.save(OUT)
print(f"✓ 生成完了: {OUT}")
print(f"  シート: {[s.title for s in wb.worksheets]}")
print(f"  データ行数: {N}件")
